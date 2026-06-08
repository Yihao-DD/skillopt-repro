"""T002 — descriptor v0 tests (behavioral fingerprint from trajectory).

Validates the Tier-A code-level descriptor (`qd/descriptor.py`):
  - φ(τ) is bounded in [0,1]^p and deterministic;
  - `code_features` parses library/control-flow/op strategy from generated code;
  - **stability (SPEC 命题 3.2)**: a skill's behavior point b is stable under
    probe-set resampling (small per-axis std) — uses the real SpreadsheetBench
    best-skill trajectory fixture;
  - **non-degeneracy (ADR-0006)**: the projection is GRADED (both axes vary and
    spread across cells) on a diverse synthetic set. The real guard is runtime
    ``n_occupied>1`` — the homogeneous SSB fixtures collapse to one cell, so a
    2-fixture best≠initial split is not assertable (see descriptor spike).

Fixtures: qd/tests/fixtures/ssb_{best,initial}_feat.jsonl (279 per-task
code-feature records each, extracted from outputs/ssb_dpsk_run1). Zero API.
"""
from __future__ import annotations

import json
import os
import random
import statistics as st

from qd.descriptor import (
    PHI_LABELS,
    code_features,
    descriptor,
    mu,
    phi,
    project,
)

_FIX = os.path.join(os.path.dirname(__file__), "fixtures")


def _load(name: str) -> list[dict]:
    with open(os.path.join(_FIX, name), encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


# ── φ bounded + deterministic ─────────────────────────────────────────────────

def test_phi_bounded_and_sized() -> None:
    for fx in ("ssb_best_feat.jsonl", "ssb_initial_feat.jsonl"):
        for t in _load(fx):
            v = phi(t)
            assert len(v) == len(PHI_LABELS) == 5
            assert all(0.0 <= x <= 1.0 for x in v), f"out-of-range φ={v}"


def test_phi_deterministic() -> None:
    t = _load("ssb_best_feat.jsonl")[0]
    assert phi(t) == phi(t)


# ── code parsing ──────────────────────────────────────────────────────────────

def test_code_features_parsing() -> None:
    code = (
        "import pandas as pd\n"
        "import openpyxl\n"
        "for i in range(3):\n"
        "    if i:\n"
        "        pd.read_excel('x')\n"
        "wb.cell(1, 1).value = 5\n"
    )
    f = code_features(code)
    assert f["uses_pandas"] == 1
    assert f["uses_openpyxl"] == 1
    assert f["n_ctrl"] == 2  # one for + one if
    assert f["n_ops"] >= 2   # read_excel + .cell(
    assert code_features("")["lines"] == 0


def test_phi_accepts_raw_code() -> None:
    v = phi({"code": "import openpyxl\nfor r in rows:\n    ws.cell(r,1)\n", "n_turns": 2})
    assert len(v) == 5 and all(0.0 <= x <= 1.0 for x in v)
    assert v[1] == 0.0  # no pandas


# ── stability under probe resampling (命题 3.2) ───────────────────────────────

def test_descriptor_stable_under_probe_resampling() -> None:
    trajs = _load("ssb_best_feat.jsonl")
    rng = random.Random(0)
    points = []
    for _ in range(30):
        sub = rng.sample(trajs, 50)  # probe set size m=50
        points.append(descriptor(sub).b)
    for axis in (0, 1):
        sd = st.pstdev(p[axis] for p in points)
        assert sd < 0.08, f"axis {axis} unstable under resampling: sd={sd:.4f}"


def test_cell_deterministic_and_in_range() -> None:
    trajs = _load("ssb_best_feat.jsonl")
    d1 = descriptor(trajs, nbins=4)
    d2 = descriptor(trajs, nbins=4)
    assert d1 == d2
    assert 0 <= d1.cell < 16


# ── separation: best vs initial (命题 3.8 dep ii) ─────────────────────────────

def test_descriptor_axes_are_graded_not_degenerate() -> None:
    # ADR-0006: real non-degeneracy guard is runtime n_occupied>1 (homogeneous
    # SSB fixtures collapse to one cell — see spike). Here we assert the
    # projection is GRADED on a diverse synthetic set: both axes vary and the set
    # spreads across multiple cells, so the projection has not degenerated.
    samples = [
        {"code": "import pandas as pd\ndf = pd.read_excel('x')\ndf.to_excel('y')\n"},
        {"code": "import openpyxl\n" + "".join(f"for c in range(3):\n    if c:\n        ws.cell({r}, c).value = {r}\n" for r in range(8))},
        {"code": "import openpyxl\nfor r in rows:\n    if r:\n        ws.cell(r, 1)\n        ws.cell(r, 2)\n"},
        {"code": "a = 1\nb = 2\nc = a + b\n"},
    ]
    pts = [descriptor([s]).b for s in samples]
    cells = {descriptor([s]).cell for s in samples}
    assert len(cells) >= 3, f"projection collapsed: cells={cells}, pts={pts}"
    for axis in (0, 1):
        spread = max(p[axis] for p in pts) - min(p[axis] for p in pts)
        assert spread > 0.1, f"axis {axis} not graded: spread={spread:.3f}"


def test_strategy_axis_tracks_op_density() -> None:
    # axis1 is op density (spreadsheet-op calls per line) — the graded strategy
    # signal that replaces the saturating 1-uses_pandas axis (ADR-0006). Both
    # samples are pandas-free, which the OLD axis mapped to an identical 1.0.
    dense = mu([{"code": "ws.cell(1, 1)\nws.cell(2, 1)\nws.cell(3, 1)\n"}])   # ~1 op / line
    sparse = mu([{"code": "import openpyxl\nx = 1\ny = 2\nz = 3\nw = 4\n"}])   # ~0 ops / line
    assert project(dense)[1] > project(sparse)[1]


def test_descriptor_spreads_real_fixtures_across_grid() -> None:
    # Real-data non-degeneracy (ADR-0006 calibration): the 558 real SpreadsheetBench
    # records must occupy most of the 16-cell grid (op-density p95-normalized →
    # 16/16; was 8/16 raw). Guards against a ref/axis change silently re-collapsing
    # the archive (the audit's P1). The old raw-op_density descriptor scores 8 here.
    recs = _load("ssb_best_feat.jsonl") + _load("ssb_initial_feat.jsonl")
    occupied = {descriptor([r]).cell for r in recs}
    assert len(occupied) >= 12, f"grid collapsed: only {len(occupied)}/16 cells occupied"
